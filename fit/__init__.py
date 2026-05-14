"""Fit performance models to benchmark results.

For each operator, fits: time_ms = a * work + b
where "work" is FLOPs (GEMM/attention matmul) or elements (norm/softmax).
Also computes implied hardware efficiency (TFLOPS / GB/s).
"""

import numpy as np
from openpyxl import load_workbook


def _get(d, *keys):
    """Return first non-None value from dict d for given keys."""
    for k in keys:
        v = d.get(k)
        if v is not None:
            return v
    return None


def load_results(path):
    """Load xlsx results, skip OOM rows, convert time_ms to float."""
    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        val = d.get("time_ms")
        if val is None or val == "OOM":
            continue
        d["time_ms"] = float(val)
        rows.append(d)
    return rows


def lstsq_fit(x, y):
    """Linear least squares: y = a * x + b. Returns (a, b, r2)."""
    x, y = np.array(x, dtype=np.float64), np.array(y, dtype=np.float64)
    mask = np.isfinite(x) & np.isfinite(y)
    x, y = x[mask], y[mask]
    if len(x) < 2:
        return 0, 0, 0
    x_mean = np.mean(x)
    x_scaled = x / x_mean
    A = np.column_stack([x_scaled, np.ones_like(x)])
    coeffs, *_ = np.linalg.lstsq(A, y, rcond=None)
    a, b = coeffs[0] / x_mean, coeffs[1]
    y_pred = a * x + b
    ss_res = float(np.sum((y - y_pred) ** 2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0
    return a, b, r2


def lstsq_log_fit(x, y):
    """Log-log fit: y = exp(c0) * x^c1. Returns (c1, exp(c0), r2)."""
    x, y = np.array(x, dtype=np.float64), np.array(y, dtype=np.float64)
    mask = (x > 0) & (y > 0) & np.isfinite(x) & np.isfinite(y)
    x, y = x[mask], y[mask]
    if len(x) < 3:
        return 1, 0, 0
    log_x = np.log(x)
    log_x_mean = np.mean(log_x)
    A = np.column_stack([log_x - log_x_mean, np.ones_like(x)])
    coeffs, *_ = np.linalg.lstsq(A, np.log(y), rcond=None)
    c1, c0_shifted = coeffs[0], coeffs[1]
    c0 = c0_shifted - c1 * log_x_mean
    y_pred = np.exp(c0) * x ** c1
    ss_res = float(np.sum((y - y_pred) ** 2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0
    return c1, np.exp(c0), r2


def fit_gemm(results):
    """Fit GEMM operators: work = M*K*N (proportional to FLOPs)."""
    print("=" * 60)
    print("GEMM Operators")
    print("=" * 60)

    gemm_ops = {"q_proj", "k_proj", "v_proj", "o_proj", "ffn_up", "ffn_gate", "ffn_down"}
    ops = sorted(set(r["op_name"] for r in results) & gemm_ops)
    all_x, all_y = [], []

    for op_name in ops:
        op_results = [r for r in results if r["op_name"] == op_name]
        work = np.array([r["M"] * r["K"] * r["N"] for r in op_results])
        time_ms = np.array([r["time_ms"] for r in op_results])

        a, b, r2 = lstsq_fit(work, time_ms)
        c1, _, r2_log = lstsq_log_fit(work, time_ms)

        flops = np.sum(2 * work)
        total_ms = np.sum(time_ms)
        avg_tflops = (flops / (total_ms / 1000)) / 1e12 if total_ms > 0 else 0

        print(f"\n{op_name}")
        print(f"  time = {a:.3e} * (M*K*N) + {b:.4f}   R2={r2:.4f}")
        print(f"  power-law exponent: {c1:.3f}   R2_log={r2_log:.4f}")
        print(f"  effective TFLOPS: {avg_tflops:.1f}")

        all_x.extend(work)
        all_y.extend(time_ms)

    a, b, r2 = lstsq_fit(all_x, all_y)
    print(f"\n--- All GEMM combined ---")
    print(f"  time = {a:.3e} * (M*K*N) + {b:.4f}   R2={r2:.4f}")


def fit_attention(results):
    """Fit attention sub-operators."""
    print("\n" + "=" * 60)
    print("Attention Operators")
    print("=" * 60)

    for op_name in ["qk_matmul", "softmax", "score_v_matmul"]:
        op_results = [r for r in results if r["op_name"] == op_name]
        if not op_results:
            continue
        b = np.array([r["b"] for r in op_results], dtype=np.float64)
        s = np.array([r["s"] for r in op_results], dtype=np.float64)
        nh = np.array([_get(r, "num_heads", "nh") for r in op_results], dtype=np.float64)
        hd = np.array([_get(r, "head_dim", "hd") for r in op_results], dtype=np.float64)
        time_ms = np.array([r["time_ms"] for r in op_results])

        if op_name == "softmax":
            work = b * nh * s * s
            desc = "b * n_heads * s^2"
        else:
            work = b * nh * s * s * hd
            desc = "b * n_heads * s^2 * d_head"

        a, b0, r2 = lstsq_fit(work, time_ms)
        print(f"\n{op_name}")
        print(f"  work = {desc}")
        print(f"  time = {a:.3e} * work + {b0:.4f}   R2={r2:.4f}")


def fit_norm(results):
    """Fit normalization operators: work = b*s*h (elements)."""
    print("\n" + "=" * 60)
    print("Norm Operators")
    print("=" * 60)

    for op_name in ["layernorm", "rmsnorm"]:
        op_results = [r for r in results if r["op_name"] == op_name]
        if not op_results:
            continue
        work = np.array([r["b"] * r["s"] * r["h"] for r in op_results])
        time_ms = np.array([r["time_ms"] for r in op_results])

        a, b, r2 = lstsq_fit(work, time_ms)
        c1, _, r2_log = lstsq_log_fit(work, time_ms)

        bw_gbps = np.sum(2 * work) / 1e9 / (np.sum(time_ms) / 1000) if np.sum(time_ms) > 0 else 0

        print(f"\n{op_name}  (work = b*s*h)")
        print(f"  time = {a:.3e} * (b*s*h) + {b:.4f}   R2={r2:.4f}")
        print(f"  power-law exponent: {c1:.3f}   R2_log={r2_log:.4f}")
        print(f"  effective bandwidth: {bw_gbps:.0f} GB/s")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Fit performance models to benchmark results")
    parser.add_argument("results_dir", nargs="?", default="results",
                        help="Directory containing xlsx result files")
    args = parser.parse_args()

    import os
    all_path = os.path.join(args.results_dir, "all_operators.xlsx")
    if not os.path.exists(all_path):
        print(f"Not found: {all_path}")
        return

    results = load_results(all_path)
    if not results:
        print("No valid results to fit.")
        return

    print(f"Loaded {len(results)} rows from {all_path}")
    fit_gemm(results)
    fit_attention(results)
    fit_norm(results)


if __name__ == "__main__":
    main()
