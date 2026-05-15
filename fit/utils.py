import numpy as np
from openpyxl import load_workbook


def _get(d, *keys):
    """Return first non-None value from dict d for given keys."""
    for k in keys:
        v = d.get(k)
        if v is not None:
            return v
    return None


def load_results(path):
    """Load xlsx results, skip OOM rows, convert time_ms to float."""
    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        val = d.get("time_ms")
        if val is None or val == "OOM":
            continue
        d["time_ms"] = float(val)
        rows.append(d)
    return rows


def lstsq_fit(x, y):
    """Linear least squares: y = a * x + b. Returns (a, b, r2)."""
    x, y = np.array(x, dtype=np.float64), np.array(y, dtype=np.float64)
    mask = np.isfinite(x) & np.isfinite(y)
    x, y = x[mask], y[mask]
    if len(x) < 2:
        return 0, 0, 0
    x_mean = np.mean(x)
    x_scaled = x / x_mean
    A = np.column_stack([x_scaled, np.ones_like(x)])
    coeffs, *_ = np.linalg.lstsq(A, y, rcond=None)
    a, b = coeffs[0] / x_mean, coeffs[1]
    y_pred = a * x + b
    ss_res = float(np.sum((y - y_pred) ** 2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0
    return a, b, r2


def lstsq_log_fit(x, y):
    """Log-log fit: y = exp(c0) * x^c1. Returns (c1, exp(c0), r2)."""
    x, y = np.array(x, dtype=np.float64), np.array(y, dtype=np.float64)
    mask = (x > 0) & (y > 0) & np.isfinite(x) & np.isfinite(y)
    x, y = x[mask], y[mask]
    if len(x) < 3:
        return 1, 0, 0
    log_x = np.log(x)
    log_x_mean = np.mean(log_x)
    A = np.column_stack([log_x - log_x_mean, np.ones_like(x)])
    coeffs, *_ = np.linalg.lstsq(A, np.log(y), rcond=None)
    c1, c0_shifted = coeffs[0], coeffs[1]
    c0 = c0_shifted - c1 * log_x_mean
    y_pred = np.exp(c0) * x ** c1
    ss_res = float(np.sum((y - y_pred) ** 2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0
    return c1, np.exp(c0), r2
